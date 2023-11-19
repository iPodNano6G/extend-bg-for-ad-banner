import os, json
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

def resize_image(image_path, target_height):
    """이미지를 높이에 맞게 비율을 유지하면서 리사이즈합니다."""
    img = Image(image_path)
    # 원본 이미지의 비율을 계산
    ratio = img.width / img.height
    # 타겟 높이에 맞게 너비를 조정
    new_width = int(ratio * target_height)
    img.width, img.height = new_width, target_height
    return img

#상대 주소를 사용하였으므로 올바른 위치에서 실행하고, image/debug/objectDetection 폴더가 있어야 한다.
cell_height = 200
wb = Workbook()
ws = wb.active
# person, text 열 길이 늘리기
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 40
# 헤더 행 추가
ws.append(["image", "person", "text"])

# 원본 이미지 파일 리스트
source_images = {f.split('_')[0]: f for f in os.listdir("../images/debug/objectDetection") if f.endswith(('.png', '.jpg', '.jpeg', '.gif'))}
# json data 불러오기
source_data = None
with open("../images/debug/objectDetection/objectDetection.json", "r") as f:
    source_data = json.load(f)

# 엑셀에 결과 삽입
row_number = 1
for product_id, source_image_name in source_images.items():
    print(row_number, product_id)
    row_number += 1
    # 이미지 리사이즈 및 삽입
    img_path = os.path.join("../images/debug/objectDetection", source_image_name)
    img = resize_image(img_path, cell_height)
    ws.row_dimensions[row_number].height = img.height
    ws.column_dimensions['A'].width = img.width * 0.13
    ws.add_image(img, 'A' + str(row_number))

    # person 결과 및 description 결과 삽입
    for image_meta in source_data:
        if image_meta["name"] == product_id:
            ws['B' + str(row_number)] = str(image_meta["person"])
            ws['C' + str(row_number)] = str(image_meta["description"])
            break

# 엑셀 파일 저장
wb.save("cloudVisionApi.xlsx")