import os, cv2
import pandas as pd
import numpy as np
from openpyxl.drawing.image import Image
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image

from skimage.feature import local_binary_pattern
from scipy.fft import fft2, fftshift

def pad_and_blur_edges(image, kernel = 0):
    """

    """
    height, width = image.shape[:2]

    new_width = height * 2

    pad_width = (new_width - width) // 2
    """
    left_pad = np.tile(image[:, :1, :], (1, pad_width, 1))
    right_pad = np.tile(image[:, -1:, :], (1, pad_width, 1))

    padded_image = np.concatenate([left_pad, image, right_pad], axis=1)"""

    padded_image = cv2.copyMakeBorder(image, 0, 0, pad_width, pad_width, cv2.BORDER_REPLICATE)

    
    #블러작업
    mask = np.zeros(padded_image.shape[:2], dtype=np.uint8)
    mask[:, :pad_width] = 1  # Left padding
    mask[:, -pad_width:] = 1  # Right padding

    blurred_image = cv2.GaussianBlur(padded_image, (kernel, kernel), 150)
    
    padded_image = np.where(mask[:, :, np.newaxis] == 1, blurred_image, padded_image)
    
    
    """
    cpyLeft = padded_image[:,:pad_width]
    #blurLeft = cv2.blur(cpyLeft, (51,11), borderType = cv2.BORDER_REFLECT_101)
    blurLeft = cv2.GaussianBlur(cpyLeft, (kernel, kernel), 150)
    padded_image[:,:pad_width] = blurLeft

    cpyRight = padded_image[:,-pad_width:]
    #blurRight = cv2.blur(cpyRight, (51,11), borderType = cv2.BORDER_REFLECT_101)
    blurRight = cv2.GaussianBlur(cpyRight, (kernel, kernel), 150)
    padded_image[:,-pad_width:] = blurRight"""
    
    return padded_image

def extract_columns(foreground_removed_image: np.ndarray, n: int, side: str) -> np.ndarray:
    """
    Extracts n columns from the specified side of the background removed image while excluding transparent pixels.

    :param background_removed_image: The image with the background removed (includes alpha channel).
    :param n: The number of columns to extract.
    :param side: The side from which to extract the columns ('left' or 'right').
    :return: The extracted columns as an ndarray.
    """
    # Check if the image has an alpha channel
    if foreground_removed_image.shape[2] != 4:
        raise ValueError("The background removed image must have an alpha channel.")

    # Determine which columns to extract
    if side == 'left':
        cols = slice(n)
    elif side == 'right':
        cols = slice(-n, None)
    else:
        raise ValueError("Side must be 'left' or 'right'.")

    # Extract the alpha channel and determine where it's not transparent
    alpha_channel = foreground_removed_image[:, :, 3]
    non_transparent_mask = alpha_channel != 0

    # Initialize an array to hold the extracted columns
    extracted_columns = np.zeros((foreground_removed_image.shape[0], n, 3), dtype=foreground_removed_image.dtype)

    # Extract the columns while excluding transparent pixels
    for i in range(3):  # Iterate over the color channels (RGB)
        channel_data = foreground_removed_image[:, :, i]
        extracted_columns[:, :, i] = np.where(non_transparent_mask[:, cols], channel_data[:, cols], 0)

    return extracted_columns

def check_opaque_edges(image):
    # 이미지를 투명도 채널과 함께 로드

    # 투명도 채널 확인 (알파 채널)
    if image.shape[2] != 4:
        raise ValueError("이미지에 투명도 채널이 없습니다.")

    alpha_channel = image[:, :, 3]

    # 좌우 가장자리에 불투명 픽셀이 있는지 확인
    left_edge = alpha_channel[:, 0]
    right_edge = alpha_channel[:, -1]

    left_touch = np.any(left_edge == 255)
    right_touch = np.any(right_edge == 255)

    return left_touch, right_touch

# Method 1: Color Diversity Analysis using Color Histogram
def analyze_color_diversity(image: np.ndarray):
    # Calculate the color histogram
    histogram = cv2.calcHist([image], [0, 1, 2], None, [256, 256, 256], [0, 256, 0, 256, 0, 256])
    
    # Normalize the histogram
    histogram_normalized = cv2.normalize(histogram, histogram).flatten()

    # Calculate the entropy
    entropy = -np.sum(histogram_normalized * np.log2(histogram_normalized + 1e-9))
    
    return entropy

# Method 2: Texture Analysis using Edge Detection and Local Binary Pattern
def analyze_texture(image: np.ndarray, threshold = 50):
    # Convert to grayscale
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    cv2.imwrite("result_gray.png", gray_image)
    # Edge Detection
    edges = cv2.Canny(gray_image, threshold, threshold)
    cv2.imwrite("result_edge.png", edges)
    edge_count = np.sum(edges > 0)
    
    # Local Binary Pattern
    lbp = local_binary_pattern(gray_image, P=8, R=1, method='uniform')
    lbp_hist, _ = np.histogram(lbp.ravel(), bins=np.arange(0, 26), range=(0, 24))
    lbp_hist_normalized = lbp_hist / np.sum(lbp_hist)
    lbp_entropy = -np.sum(lbp_hist_normalized * np.log2(lbp_hist_normalized + 1e-9))
    
    return edge_count, lbp_entropy

# Method 3: Frequency Domain Analysis using Fourier Transform
def analyze_frequency_domain(image: np.ndarray):
    # Convert to grayscale
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # Apply Fourier Transform
    f_transform = fft2(gray_image)
    f_shifted = fftshift(f_transform)
    f_magnitude = 20 * np.log(np.abs(f_shifted))
    
    # Calculate the amount of high frequency components
    mean_frequency_amplitude = np.mean(f_magnitude)
    
    return mean_frequency_amplitude

def create_bg_only_image(original_image, background_removed_image):
    if original_image.shape[2] == 3:
        # Convert 3-channel image to 4-channel
        original_image = cv2.cvtColor(original_image, cv2.COLOR_BGR2BGRA)
    if background_removed_image.shape[2] == 3:
        # Convert 3-channel image to 4-channel
        background_removed_image = cv2.cvtColor(background_removed_image, cv2.COLOR_BGR2BGRA)
        
    alpha_channel = background_removed_image[:, :, 3]
    mask = (alpha_channel < 128)
    bg_only_image = np.zeros_like(original_image)
    for i in range(3):
        bg_only_image[:, :, i] = np.where(mask, original_image[:, :, i], 0)
    bg_only_image[:, :, 3] = np.where(mask, 255, 0)

    return bg_only_image

def resize_image(img, target_height):
    """이미지를 높이에 맞게 비율을 유지하면서 리사이즈합니다."""
    # 원본 이미지의 비율을 계산
    ratio = img.width / img.height
    # 타겟 높이에 맞게 너비를 조정
    new_width = int(ratio * target_height)
    img.width, img.height = new_width, target_height
    return img




def process_images_to_excel2(original_dir, background_removed_dir, excel_path, edge_threshold = 12, kernel = 0):
    # 워크북과 워크시트 생성
    wb = Workbook()
    ws = wb.active

    # 엑셀 헤더 설정
    ws.append([
        'product_id', 'original_image', 'background_removed_image','edge_count_left', 'edge_count_right'
    ])

    # List all files in the original images directory
    foreground_removed_image_folder = os.path.join(original_dir, "bg_only")
    if not os.path.isdir(foreground_removed_image_folder):
        os.mkdir(foreground_removed_image_folder)
    for filename in os.listdir(original_dir):
        if filename.endswith(".jpg"):
            product_id = os.path.splitext(filename)[0]
            original_image_path = os.path.join(original_dir, filename)

            background_removed_image_filename = f"{product_id}.png"
            background_removed_image_path = os.path.join(background_removed_dir, background_removed_image_filename)

            original_image = cv2.imread(original_image_path)
            background_removed_image = cv2.imread(background_removed_image_path, cv2.IMREAD_UNCHANGED)
            left, right = check_opaque_edges(background_removed_image)

            if left or right:
                continue
            else:
                foreground_removed_image = create_bg_only_image(original_image, background_removed_image)
                extract_removed_image_left = extract_columns(foreground_removed_image, 20, "left")
                extract_removed_image_right = extract_columns(foreground_removed_image, 20, "right")
                #분석 수행
                #entropy = analyze_color_diversity(extract_removed_image)
                edge_count_left, _ = analyze_texture(extract_removed_image_left)
                edge_count_right, _ = analyze_texture(extract_removed_image_right)

                # 이미지를 엑셀에 삽입
                img_original = Image(os.path.join(original_dir, filename))
                img_original = resize_image(img_original, 300)
                img_original.anchor = 'B' + str(ws.max_row + 1)  # 예를 들어 B2 셀에 배치
                ws.add_image(img_original)

                if edge_count_left <= edge_threshold and edge_count_right <= edge_threshold:
                    # 피사체가 제거된 이미지를 리사이즈하여 배치
                    background_removed_image_filename = f"{product_id}.png"
                    foreground_removed_image_path = os.path.join(foreground_removed_image_folder ,background_removed_image_filename)
                    print(foreground_removed_image_path)
                    #cv2.imwrite(foreground_removed_image_path, foreground_removed_image)
                    cv2.imwrite(foreground_removed_image_path, pad_and_blur_edges(original_image, kernel))
                    img_foreground_removed = Image(foreground_removed_image_path)
                    img_foreground_removed = resize_image(img_foreground_removed, 300)
                    img_foreground_removed.anchor = 'C' + str(ws.max_row + 1)  # 예를 들어 C2 셀에 배치
                    ws.add_image(img_foreground_removed)

            # 결과 데이터 추가
            ws.append([
                product_id,
                '',  # 이미지는 이미 삽입되었으므로 경로 대신 빈 문자열을 사용
                '',  # 이미지는 이미 삽입되었으므로 경로 대신 빈 문자열을 사용
                edge_count_left, edge_count_right
            ])

    # Excel 파일 저장
    wb.save(excel_path)
# 사용 예
kernel = 41
process_images_to_excel2(r'C:\projects\extend-bg-for-ad-banner\extract\border_removed_result',
         r"C:\projects\extend-bg-for-ad-banner\extract\border_removed_result\ps_background_removal2",
         "2023_11_20_extend_border_50_12_gausian_blur_" + str(kernel) + "_" +str(kernel)+ ".xlsx", kernel = kernel)


"""path = r"C:\projects\extend-bg-for-ad-banner\extract\border_removed_result\1AJA2D403BK.jpg"
bg_path = r"C:\projects\extend-bg-for-ad-banner\extract\border_removed_result\ps_background_removal2\1AJA2D403BK.png"
img = cv2.imread(path, cv2.IMREAD_UNCHANGED)


padded_image = pad_and_blur_edges(img)
cv2.imwrite("result2.png", padded_image)
bg_removed_img = cv2.imread(bg_path, cv2.IMREAD_UNCHANGED)
l,r = check_opaque_edges(bg_removed_img)
print(l,r)
print(path)
#img = cv2.cvtColor(img ,cv2.COLOR_BGR2BGRA)
img = create_bg_only_image(img, bg_removed_img)

extract_removed_image = extract_columns(img, 20, "left")
print(extract_removed_image.shape)

#cv2.imwrite("result2.png",extract_removed_image)
edge_count, lbp_entropy = analyze_texture(extract_removed_image)
print(edge_count, lbp_entropy)"""