import os, json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

def resize_image(image_path, target_height):
    """이미지를 높이에 맞게 비율을 유지하면서 리사이즈합니다."""
    img = Image(image_path)
    # 원본 이미지의 비율을 계산
    ratio = img.width / img.height
    # 타겟 높이에 맞게 너비를 조정
    new_width = int(ratio * target_height)
    img.width, img.height = new_width, target_height
    return img

def make_xls(source_folder, result_folder, cell_height=200):
    # 워크북 생성
    wb = Workbook()
    ws = wb.active

    # 엑셀의 행 높이 설정 (단위: 20은 약 1 포인트)
    ws.row_dimensions[1].height = cell_height / 0.75  # 엑셀의 행 높이는 포인트 단위로 설정됨

    # 헤더 행 추가
    ws.append(["원본 이미지", "결과 이미지", "배점칸"])

    # 원본 이미지 파일 리스트
    source_images = {f.split('.')[0]: f for f in os.listdir(source_folder) if f.endswith(('.png', '.jpg', '.jpeg', '.gif'))}
    # 결과 이미지 파일 리스트 (product_id를 키로 함)
    result_images = {f.split('_output')[0]: f for f in os.listdir(result_folder) if '_output' in f}

    # 이미지를 엑셀에 삽입
    for product_id, source_image_name in source_images.items():
        row_number = ws.max_row + 1
        ws.row_dimensions[row_number].height = cell_height

        # 원본 이미지 리사이즈 및 삽입
        img_path = os.path.join(source_folder, source_image_name)
        img = resize_image(img_path, cell_height)
        ws.add_image(img, 'A' + str(row_number))

        # 결과 이미지 리사이즈 및 삽입
        result_image_name = result_images.get(product_id)
        if result_image_name:
            img_path = os.path.join(result_folder, result_image_name)
            img = resize_image(img_path, cell_height)
            ws.add_image(img, 'B' + str(row_number))


        # 배점칸은 비워둠
        ws['C' + str(row_number)] = ""
    # 셀 너비 설정
    ws.column_dimensions['A'].width = 15  # 엑셀의 너비는 캐릭터 크기 단위로 설정됨
    ws.column_dimensions['B'].width = 40  # 엑셀의 너비는 캐릭터 크기 단위로 설정됨

    # 엑셀 파일 저장
    now = datetime.now()

    wb.save(f"result_{now.month}-{now.day}.xlsx")

def make_xls_modified_for_simple_fill(source_folder, result_folder1, result_folder2, cell_height=200):
    # 워크북 생성
    wb = Workbook()
    ws = wb.active

    # 엑셀의 행 높이 설정 (단위: 20은 약 1 포인트)
    ws.row_dimensions[1].height = cell_height / 0.9  # 엑셀의 행 높이는 포인트 단위로 설정됨

    # 헤더 행 추가
    ws.append(["원본 이미지", "결과 이미지1", "결과 이미지2", "색상 편차 그래프", "배점칸"])

    # 원본 이미지 파일 리스트
    source_images = {f.split('.')[0]: f for f in os.listdir(source_folder) if f.endswith(('.png', '.jpg', '.jpeg', '.gif'))}
    # 결과 이미지 파일 리스트 (product_id를 키로 함)
    result_images = {f.split('_result')[0]: f for f in os.listdir(result_folder1) if '_result' in f}
    #result_gray = {f.split('_gray')[0]: f for f in os.listdir(result_folder1) if '_gray' in f}
    #result_graph = {f.split('_graph')[0]: f for f in os.listdir(result_folder1) if '_graph' in f}
    #result_diff = {f.split('_diff')[0]: f for f in os.listdir(result_folder1) if '_diff' in f}
    divided_images = {f.split('_div')[0]: f for f in os.listdir(result_folder1) if '_div' in f}
    with open(result_folder1+"/simpleInfo.json", "r") as f:
        simpleInfo = json.load(f)

    # 이미지를 엑셀에 삽입
    for product_id, source_image_name in source_images.items():
        row_number = ws.max_row + 1
        ws.row_dimensions[row_number].height = cell_height

        # 원본 이미지 리사이즈 및 삽입
        img_path = os.path.join(source_folder, source_image_name)
        img = resize_image(img_path, cell_height)
        ws.add_image(img, 'A' + str(row_number))

        # 결과 이미지 리사이즈 및 삽입
        # result_gray_name = result_gray.get(product_id)
        # if result_gray_name:
        #     img_path = os.path.join(result_folder1, result_gray_name)
        #     img = resize_image(img_path, cell_height)
        #     ws.add_image(img, 'B' + str(row_number))

        result_image_name = result_images.get(product_id)
        if result_image_name:
            img_path = os.path.join(result_folder1, result_image_name)
            img = resize_image(img_path, cell_height)
            ws.add_image(img, 'B' + str(row_number))
            
        divided_image_name = divided_images.get(product_id)
        if divided_image_name:
            img_path = os.path.join(result_folder1, divided_image_name)
            img = resize_image(img_path, cell_height)
            ws.add_image(img, 'C' + str(row_number))

        # result_graph_name = result_graph.get(product_id)
        # if result_graph_name:
        #     img_path = os.path.join(result_folder1, result_graph_name)
        #     img = resize_image(img_path, cell_height)
        #     ws.add_image(img, 'D' + str(row_number))

        # result_diff_name = result_diff.get(product_id)
        # if result_diff_name:
        #     img_path = os.path.join(result_folder1, result_diff_name)
        #     img = resize_image(img_path, cell_height)
        #     ws.add_image(img, 'E' + str(row_number))

        # 배점칸은 비워둠
        # ws['F' + str(row_number)] = simpleInfo[product_id].get('isSimple')
        # ws['G' + str(row_number)] = simpleInfo[product_id].get('dtwDistance')
        ws['D' + str(row_number)] = simpleInfo[product_id].get('isSimple')
        fill_history = simpleInfo[product_id].get('fillHistory')
        ws['E' + str(row_number)] = "/".join([','.join([str(c) for c in lst]) for lst in fill_history])
    # 셀 너비 설정
    ws.column_dimensions['A'].width = 30  # 엑셀의 너비는 캐릭터 크기 단위로 설정됨
    ws.column_dimensions['B'].width = 30  # 엑셀의 너비는 캐릭터 크기 단위로 설정됨
    ws.column_dimensions['C'].width = 30  # 엑셀의 너비는 캐릭터 크기 단위로 설정됨
    ws.column_dimensions['D'].width = 15  # 엑셀의 너비는 캐릭터 크기 단위로 설정됨
    ws.column_dimensions['E'].width = 30  # 엑셀의 너비는 캐릭터 크기 단위로 설정됨


    # 엑셀 파일 저장
    now = datetime.now()

    wb.save(f"result_{now.month}-{now.day}_final.xlsx")

# 사용 예
# make_xls(r'C:\projects\extend-bg-for-ad-banner\extract', r'C:\projects\extend-bg-for-ad-banner\extract\border_removed_result\2023_11_06_test', 200)
# xls 폴더에서 실행하면 debug
make_xls_modified_for_simple_fill("../images", "../images/simpleNoblurred", "../images/simpleBlurred", 300)
